#!/usr/bin/env python3
"""
Generate a nightly Excel performance report from Qwen-Image diffusion benchmark JSON results.

Reads diffusion_perf_*.json files from tests/perf/results/ (or DIFFUSION_BENCHMARK_DIR),
groups by test_name, highlights cells where metrics changed vs previous run,
and outputs a .xlsx report.

Usage:
    python tools/nightly/generate_qwen_image_perf_excel.py
    python tools/nightly/generate_qwen_image_perf_excel.py \
        --input-dir tests/perf/results \
        --output-file tests/qwen_image_perf_report.xlsx
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
from collections.abc import Iterable, Sequence
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

LOGGER = logging.getLogger(__name__)

GREY_BLOCK_FILL = PatternFill(start_color="D3D3D3", fill_type="solid")

# Metrics that get float coercion and number formatting in Excel
BENCHMARK_COLUMNS: tuple[str, ...] = (
    "duration",
    "completed_requests",
    "failed_requests",
    "throughput_qps",
    "latency_mean",
    "latency_median",
    "latency_p50",
    "latency_p99",
    "peak_memory_mb_max",
    "peak_memory_mb_mean",
    "peak_memory_mb_median",
    "slo_attainment_rate",
)

NUMERIC_FORMAT_COLUMNS: tuple[str, ...] = BENCHMARK_COLUMNS

SUMMARY_COLUMNS: tuple[str, ...] = (
    "date",
    "test_name",
    "model",
    "backend",
    "dataset",
    "task",
    "completed_requests",
    "failed_requests",
    "duration",
    "throughput_qps",
    "latency_mean",
    "latency_median",
    "latency_p50",
    "latency_p99",
    "peak_memory_mb_max",
    "peak_memory_mb_mean",
    "peak_memory_mb_median",
    "slo_attainment_rate",
    "slo_met_success",
    "slo_scale",
    "commit_sha",
    "build_id",
    "build_url",
    "source_file",
)

# Regex to extract date from filename: diffusion_perf_<test_name>_<YYYYMMDD-HHMMSS>.json
_FILENAME_DATE_RE = re.compile(r"_(\d{8}-\d{6})\.json$")

_DEFAULT_RESULT_DIR = Path(__file__).parent.parent.parent / "tests" / "perf" / "results"
DEFAULT_INPUT_DIR = os.getenv("DIFFUSION_BENCHMARK_DIR", str(_DEFAULT_RESULT_DIR))
DEFAULT_OUTPUT_DIR = os.getenv("DEFAULT_OUTPUT_DIR", str(Path(__file__).parent.parent.parent / "tests"))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Read diffusion_perf_*.json files and generate a Qwen-Image Excel perf report."
    )
    ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    parser.add_argument(
        "--input-dir",
        default=DEFAULT_INPUT_DIR,
        help="Directory containing diffusion_perf_*.json files.",
    )
    parser.add_argument(
        "--output-file",
        default=os.path.join(DEFAULT_OUTPUT_DIR, f"qwen_image_perf_{ts}.xlsx"),
        help="Output path of the Excel report.",
    )
    parser.add_argument("--commit-sha", default=None)
    parser.add_argument("--build-id", default=None)
    parser.add_argument("--build-url", default=None)
    return parser.parse_args()


def _extract_date_from_filename(filename: str) -> str:
    """Extract YYYYMMDD-HHMMSS from diffusion_perf_<name>_<date>.json, fallback to now."""
    m = _FILENAME_DATE_RE.search(filename)
    if m:
        return m.group(1)
    return datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")


def _iter_json_records(input_dir: str) -> Iterable[dict[str, Any]]:
    """Iterate over diffusion_perf_*.json files and yield normalized records."""
    dir_path = Path(input_dir)
    if not dir_path.is_dir():
        LOGGER.warning("input dir '%s' does not exist", input_dir)
        return

    for entry in sorted(dir_path.iterdir()):
        if not entry.is_file():
            continue
        # Accept all .json files (not just diffusion_perf_ prefix) for flexibility
        if not entry.suffix == ".json":
            continue

        try:
            with open(entry, encoding="utf-8") as f:
                data = json.load(f)
        except (OSError, json.JSONDecodeError) as exc:
            LOGGER.warning("failed to load '%s': %s", entry, exc)
            continue

        if not isinstance(data, dict):
            LOGGER.warning("'%s' root is not an object, skip", entry)
            continue

        record: dict[str, Any] = dict(data)
        record.setdefault("date", _extract_date_from_filename(entry.name))

        # Extract test_name from filename if not in JSON
        if "test_name" not in record:
            # diffusion_perf_<test_name>_<timestamp>.json
            stem = entry.stem  # e.g. diffusion_perf_test_qwen_image_ulysses2_cfg2_20250101-120000
            m = re.match(r"diffusion_perf_(.+)_\d{8}-\d{6}$", stem)
            record["test_name"] = m.group(1) if m else stem

        record["source_file"] = entry.name
        yield record


def _collect_records(input_dir: str) -> list[dict[str, Any]]:
    return list(_iter_json_records(input_dir))


def _apply_build_metadata_to_latest_only(
    records: Sequence[dict[str, Any]],
    commit_sha: str | None,
    build_id: str | None,
    build_url: str | None,
) -> None:
    """Attach build metadata only to rows with the latest date."""
    if not records:
        return
    max_date = max((r.get("date") or "") for r in records)
    for r in records:
        if (r.get("date") or "") == max_date:
            r["commit_sha"] = commit_sha
            r["build_id"] = build_id
            r["build_url"] = build_url
        else:
            r["commit_sha"] = None
            r["build_id"] = None
            r["build_url"] = None


def _sort_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Group by test_name, newest date first within each group."""
    by_date_desc = sorted(records, key=lambda r: (r.get("date") or ""), reverse=True)
    return sorted(by_date_desc, key=lambda r: (r.get("test_name") or ""))


def _values_differ(a: Any, b: Any) -> bool:
    if a is None and b is None:
        return False
    if a is None or b is None:
        return True
    if isinstance(a, float) and isinstance(b, float):
        if a != a and b != b:
            return False
        if a != a or b != b:
            return True
        return abs(a - b) > 1e-9
    return a != b


def _apply_change_highlight(
    ws,
    columns: Sequence[str],
    records: Sequence[dict[str, Any]],
) -> None:
    """Grey cells in the latest row of each test_name when a benchmark metric changed."""
    if not records:
        return
    col_index = {c: i + 1 for i, c in enumerate(columns)}
    i = 0
    while i < len(records):
        test_name = records[i].get("test_name")
        block_start = i
        while i < len(records) and records[i].get("test_name") == test_name:
            i += 1
        if block_start + 1 >= i:
            continue  # only one row for this test_name, nothing to compare
        newest_idx = block_start
        prev_idx = block_start + 1
        excel_row = newest_idx + 2  # +1 for header, +1 for 1-indexed
        for col in BENCHMARK_COLUMNS:
            if col not in col_index:
                continue
            cur = records[newest_idx].get(col)
            prev = records[prev_idx].get(col)
            if _values_differ(cur, prev):
                ws.cell(row=excel_row, column=col_index[col]).fill = GREY_BLOCK_FILL


def _to_float_if_numeric(value: Any) -> Any:
    if value is None:
        return value
    if isinstance(value, (int, float)):
        return float(value) if isinstance(value, int) else value
    if isinstance(value, str):
        try:
            return float(value)
        except (ValueError, TypeError):
            return value
    return value


def _write_sheet(
    ws,
    columns: Sequence[str],
    rows: Iterable[dict[str, Any]],
    numeric_columns: Sequence[str] = (),
) -> None:
    numeric_set = set(numeric_columns)
    ws.append(list(columns))
    for record in rows:
        row_values = []
        for col in columns:
            v = record.get(col)
            if col in numeric_set:
                v = _to_float_if_numeric(v)
            row_values.append(v)
        ws.append(row_values)


def _format_numeric_columns(ws, columns: Sequence[str], num_rows: int) -> None:
    numeric_set = set(NUMERIC_FORMAT_COLUMNS)
    for c, col_name in enumerate(columns):
        if col_name not in numeric_set:
            continue
        col_letter = get_column_letter(c + 1)
        ws.column_dimensions[col_letter].width = 16
        for r in range(2, 2 + num_rows):
            cell = ws.cell(row=r, column=c + 1)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0000"
            elif isinstance(cell.value, str):
                try:
                    cell.value = float(cell.value)
                    cell.number_format = "0.0000"
                except (ValueError, TypeError):
                    pass


def _build_raw_columns(records: Sequence[dict[str, Any]], summary_cols: Sequence[str]) -> list[str]:
    keys: set[str] = set()
    for r in records:
        keys.update(r.keys())
    ordered: list[str] = []
    for k in summary_cols:
        if k in keys:
            ordered.append(k)
            keys.discard(k)
    ordered.extend(sorted(keys))
    return ordered


def generate_excel_report(
    input_dir: str,
    output_file: str,
    commit_sha: str | None,
    build_id: str | None,
    build_url: str | None,
) -> None:
    records = _collect_records(input_dir)
    if not records:
        LOGGER.warning("no valid json records found under '%s'", input_dir)

    sorted_records = _sort_records(records)
    _apply_build_metadata_to_latest_only(sorted_records, commit_sha, build_id, build_url)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "summary"

    _write_sheet(ws_summary, SUMMARY_COLUMNS, sorted_records, numeric_columns=NUMERIC_FORMAT_COLUMNS)
    _format_numeric_columns(ws_summary, SUMMARY_COLUMNS, len(sorted_records))
    _apply_change_highlight(ws_summary, SUMMARY_COLUMNS, sorted_records)

    if sorted_records:
        raw_columns = _build_raw_columns(sorted_records, SUMMARY_COLUMNS)
        ws_raw = wb.create_sheet(title="raw")
        _write_sheet(ws_raw, raw_columns, sorted_records)

    Path(output_file).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    LOGGER.info("excel report saved to '%s'", output_file)


def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )
    args = parse_args()
    commit_sha = args.commit_sha or os.getenv("BUILDKITE_COMMIT")
    build_id = args.build_id or os.getenv("BUILDKITE_BUILD_ID")
    build_url = args.build_url or os.getenv("BUILDKITE_BUILD_URL")

    generate_excel_report(
        input_dir=args.input_dir,
        output_file=args.output_file,
        commit_sha=commit_sha,
        build_id=build_id,
        build_url=build_url,
    )


if __name__ == "__main__":
    main()
